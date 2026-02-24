# 📚 GUÍA RÁPIDA — Todo Lo Extraído de las 6 Páginas
#
# Este archivo es una CHEATSHEET de referencia rápida con los patterns
# y snippets más útiles extraídos de cada fuente. Para el detalle completo
# ver DEEP_DIVE_COMPLETO.md

# ============================================================================
# 1. CURL_CFFI — Cheatsheet Completo
# Fuente: https://curl-cffi.readthedocs.io/en/latest/
# ============================================================================

"""
INSTALACIÓN:
    pip install curl_cffi --upgrade

GET BÁSICO CON IMPERSONACIÓN:
    import curl_cffi
    r = curl_cffi.get("https://www.fravega.com", impersonate="chrome")

SESSION CON COOKIES (SIEMPRE USAR):
    with curl_cffi.Session(impersonate="chrome") as s:
        s.get("https://www.fravega.com")  # Primera visita, guarda cookies
        r = s.get("https://www.fravega.com/api/graphql")  # Usa cookies

POST JSON (GRAPHQL):
    payload = {"query": "{ search(term: 'iphone') { products { name price } } }"}
    r = curl_cffi.post("https://www.fravega.com/api/graphql", json=payload, impersonate="chrome")

RETRY NATIVO:
    from curl_cffi import Session, RetryStrategy
    strategy = RetryStrategy(count=3, delay=0.2, jitter=0.1, backoff="exponential")
    with Session(impersonate="chrome", retry=strategy) as s:
        r = s.get(url)

ASYNC PARALELO:
    from curl_cffi import AsyncSession
    async with AsyncSession(impersonate="chrome") as s:
        tasks = [s.get(url) for url in urls]
        results = await asyncio.gather(*tasks)

HTTP/3 (MENOS DETECCIÓN):
    r = curl_cffi.get(url, http_version="v3")

PROXY:
    r = curl_cffi.get(url, proxy="http://user:pass@proxy:3128", impersonate="chrome")

VERIFICAR FINGERPRINT:
    r = curl_cffi.get("https://tls.browserleaks.com/json", impersonate="chrome")
    print(r.json()["ja3n_hash"])

HEADERS CUSTOM (SE SUMAN A LOS DE CHROME):
    r = curl_cffi.get(url, impersonate="chrome", headers={"Accept-Language": "es-AR"})

DESACTIVAR HEADERS DEFAULT:
    r = curl_cffi.get(url, impersonate="chrome", default_headers=False, headers={...})

WEBSOCKETS:
    from curl_cffi import Session, WebSocket
    def on_message(ws: WebSocket, message): print(message)
    with Session() as s:
        ws = s.ws_connect("wss://api.example.com/prices", on_message=on_message)
        ws.run_forever()

UPLOAD ARCHIVOS:
    mp = curl_cffi.CurlMime()
    mp.addpart(name="file", filename="data.xlsx", local_path="./data.xlsx")
    r = curl_cffi.post(url, multipart=mp)

KEEP ALIVE HTTP/2:
    s = Session(impersonate="chrome")
    s.get(url)
    s.upkeep()  # Ping frame para mantener conexión

DESCARTAR COOKIES:
    s = curl_cffi.Session(discard_cookies=True)

FORZAR HTTP/1.1 (SI H2 FALLA):
    from curl_cffi import CurlHttpVersion
    r = curl_cffi.get(url, http_version=CurlHttpVersion.V1_1)
"""


# ============================================================================
# 2. CIRCUIT BREAKER — Pattern de Resiliencia
# Fuente: https://skills.sh/wshobson/agents/error-handling-patterns
# ============================================================================

"""
ESTADOS:
    CLOSED    → Normal, requests pasan
    OPEN      → API caída, rechaza inmediatamente (ahorra tiempo y recursos)
    HALF_OPEN → Prueba con 1 request, si funciona → CLOSED, si falla → OPEN

CUÁNDO USAR:
    - API de Frávega empieza a dar 403 → abrir circuito, no bombardear
    - Servidor da 500 repetidamente → abrir, esperar recovery
    - Rate limit (429) → abrir, esperar cooldown

IMPLEMENTACIÓN: ver core/http_client.py → CircuitBreaker class
"""


# ============================================================================
# 3. ASYNC PATTERNS — Scraping Paralelo
# Fuente: https://skills.sh/wshobson/agents/async-python-patterns
# ============================================================================

"""
SEMAPHORE RATE LIMITING:
    semaphore = asyncio.Semaphore(3)  # Max 3 requests paralelos
    async with semaphore:
        response = await session.get(url)

BATCH PROCESSING:
    async def process_batch(urls, batch_size=10):
        for i in range(0, len(urls), batch_size):
            batch = urls[i:i+batch_size]
            results = await asyncio.gather(*[fetch(url) for url in batch])

TIMEOUT:
    result = await asyncio.wait_for(operation, timeout=10.0)

ERROR HANDLING EN GATHER:
    results = await asyncio.gather(*tasks, return_exceptions=True)
    for r in results:
        if isinstance(r, Exception):
            logger.error(f"Task falló: {r}")
        else:
            process(r)

PRODUCER-CONSUMER:
    queue = asyncio.Queue(maxsize=100)
    # Producer pone URLs en la queue
    # Consumer las procesa con rate limiting
"""


# ============================================================================
# 4. SYSTEMATIC DEBUGGING — Proceso de 4 Fases
# Fuente: https://skills.sh/obra/superpowers/systematic-debugging
# ============================================================================

"""
IRON LAW: NO FIXES WITHOUT ROOT CAUSE INVESTIGATION FIRST

FASE 1 - ROOT CAUSE:
    1. Leer el error completo (stack trace, logs)
    2. Identificar el componente exacto que falla
    3. Agregar logs en cada frontera (Multi-Layer Diagnostic)

FASE 2 - PATTERN ANALYSIS:
    1. ¿Es reproducible? ¿Intermitente?
    2. ¿Cuándo empezó? ¿Qué cambió?
    3. ¿Data Flow: de dónde viene el valor malo?

FASE 3 - HYPOTHESIS + TEST:
    1. Formular hipótesis: "El 403 es por fingerprint TLS"
    2. Diseñar test: "comparar JA3 de curl_cffi vs Chrome real"
    3. Ejecutar y verificar

FASE 4 - IMPLEMENTATION:
    1. Fix mínimo que resuelve la root cause
    2. Agregar test que reproduzca el bug
    3. Verificar que pasa

REGLA DE 3: Si 3 fixes fallan → cuestionar la ARQUITECTURA
"""


# ============================================================================
# 5. MCP SERVER — Template para Price Monitor
# Fuente: https://modelcontextprotocol.io/docs/develop/build-server
# ============================================================================

"""
SETUP:
    pip install "mcp[cli]" httpx

TEMPLATE MCP SERVER:
    from mcp.server.fastmcp import FastMCP
    mcp = FastMCP("price-monitor")
    
    @mcp.tool()
    async def get_prices(product: str) -> str:
        '''Get prices for a product across stores.'''
        # Tu lógica de scraping aquí
        return json.dumps(prices)
    
    @mcp.tool()
    async def find_deals(category: str, min_discount: float = 20.0) -> str:
        '''Find deals with minimum discount percentage.'''
        return json.dumps(deals)
    
    def main():
        mcp.run(transport="stdio")

LOGGING (NUNCA print()):
    import sys
    print("debug info", file=sys.stderr)  # ✅
    logging.info("debug info")            # ✅
    print("debug info")                   # ❌ Rompe STDIO transport
"""


# ============================================================================
# 6. CLAUDE CODE — Skills y Subagents Custom
# Fuente: https://docs.anthropic.com/en/docs/claude-code/best-practices
# ============================================================================

"""
SKILL CUSTOM (guardar en .claude/skills/nombre/SKILL.md):
    ---
    name: scrape-target
    description: Scrape a new e-commerce target
    ---
    1. Analyze target URL structure
    2. Find API endpoints (GraphQL, REST)
    3. Create scraper with curl_cffi
    4. Add rate limiting + error handling
    5. Write tests
    6. Add to monitoring rotation
    
    USO: /scrape-target https://www.garbarino.com

SUBAGENT CUSTOM (guardar en .claude/agents/nombre.md):
    ---
    name: price-analyzer
    description: Analyzes price data for arbitrage
    tools: Read, Grep, Glob, Bash
    ---
    You are a price analysis specialist. Given product data:
    - Identify lowest prices across stores
    - Calculate potential margins (>15% = interesting)
    - Flag suspicious pricing
    - Generate comparison report

HEADLESS MODE:
    claude -p "Scrape Frávega celulares and save to DB" --output-format json
    
FAN OUT (procesar múltiples targets):
    for target in fravega garbarino musimundo; do
        claude -p "Scrape $target" --allowedTools "Edit,Bash(python *)"
    done

SESSION MANAGEMENT:
    /clear          → Limpiar contexto entre tareas
    /compact        → Comprimir contexto (mantiene lo importante)
    /rewind         → Deshacer cambios
    claude --continue → Retomar última conversación
"""


# ============================================================================
# 7. EXCEL CON FÓRMULAS — Reportes Dinámicos
# Fuente: https://skills.sh/anthropics/skills/xlsx
# ============================================================================

"""
REGLA DE ORO: Usar fórmulas Excel, NUNCA hardcodear valores

LEER EXCEL:
    import pandas as pd
    df = pd.read_excel("Precios competidores.xlsx", engine="openpyxl")

CREAR EXCEL CON FÓRMULAS:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Producto"
    ws["B1"] = "Precio Frávega"
    ws["C1"] = "Precio Garbarino"
    ws["D1"] = "Diferencia"
    ws["D2"] = "=C2-B2"           # ← FÓRMULA, no valor
    ws["E2"] = "=D2/B2*100"      # ← % de diferencia
    wb.save("comparacion.xlsx")

COLOR CODING:
    from openpyxl.styles import Font
    ws["B2"].font = Font(color="0000FF")  # Azul = inputs
    ws["D2"].font = Font(color="000000")  # Negro = fórmulas
    ws["E2"].font = Font(color="FF0000")  # Rojo = alertas

LEER SIN PERDER FÓRMULAS:
    wb = openpyxl.load_workbook("file.xlsx")  # Mantiene fórmulas
    # ⚠️ data_only=True lee valores pero PIERDE fórmulas al guardar

ARCHIVOS GRANDES:
    wb = openpyxl.load_workbook("big.xlsx", read_only=True)
"""


# ============================================================================
# 8. RETRY + EXPONENTIAL BACKOFF — Pattern
# Fuente: error-handling-patterns skill
# ============================================================================

"""
import time
import random

def retry_with_backoff(func, max_attempts=3, base_delay=1.0, backoff=2.0):
    for attempt in range(max_attempts):
        try:
            return func()
        except Exception as e:
            if attempt == max_attempts - 1:
                raise
            delay = base_delay * (backoff ** attempt) + random.uniform(0, 0.5)
            print(f"Intento {attempt+1} falló: {e}. Retry en {delay:.1f}s...")
            time.sleep(delay)

# curl_cffi tiene esto BUILT-IN:
from curl_cffi import RetryStrategy
strategy = RetryStrategy(count=3, delay=0.5, jitter=0.2, backoff="exponential")
"""


# ============================================================================
# 9. GRACEFUL DEGRADATION — Múltiples Fallbacks
# Fuente: error-handling-patterns skill
# ============================================================================

"""
def get_price_with_fallback(product_id):
    '''Intentar múltiples fuentes hasta que una funcione.'''
    
    # Intento 1: API en vivo
    try:
        return fetch_live_api(product_id)
    except Exception:
        pass
    
    # Intento 2: Cache local
    try:
        return read_from_cache(product_id)
    except Exception:
        pass
    
    # Intento 3: Último precio conocido en DB
    try:
        return db.get_latest_price(product_id)
    except Exception:
        pass
    
    # Fallback final
    return {"price": None, "source": "unavailable", "stale": True}
"""


# ============================================================================
# 10. TDD — Test-Driven Development
# Fuente: https://skills.sh/obra/superpowers/test-driven-development
# ============================================================================

"""
CICLO: RED → GREEN → REFACTOR

1. RED: Escribir test que FALLA
    def test_detect_glitch_extreme_discount():
        product = Product(name="iPhone", current_price=100, list_price=1000)
        glitch = sniffer.detect_glitch(product)
        assert glitch is not None
        assert glitch.severity == "high"

2. GREEN: Código MÍNIMO para que pase
    def detect_glitch(self, product):
        discount = product.calculated_discount
        if discount >= 40:
            return Glitch(product=product, severity="high", ...)

3. REFACTOR: Mejorar sin romper tests
    # Separar heurísticas, agregar más checks, etc.

BUG FIX: SIEMPRE escribir test que reproduzca el bug PRIMERO
    def test_bug_403_on_graphql():
        # Este test debe FALLAR antes del fix
        response = client.post(graphql_url, json=query)
        assert response.status_code != 403
"""
